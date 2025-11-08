from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
from firebase_config import get_db, get_all_documents, add_document, get_document
from timetable_generator import TimetableGenerator
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import io
import json
from datetime import datetime
import uuid

app = Flask(__name__)
CORS(app)  

db = get_db()

@app.route('/api/health', methods=['GET'])
def health_check():
    return jsonify({'status': 'healthy', 'message': 'Timetable Generator API is running'})


@app.route('/api/courses', methods=['GET', 'POST'])
def manage_courses():
    """Get all courses or add a new course"""
    if request.method == 'GET':
        result = get_all_documents('courses')
        return jsonify(result)
    
    elif request.method == 'POST':
        data = request.json
        course_id = data.get('id', str(uuid.uuid4()))
        result = add_document('courses', course_id, data)
        return jsonify(result)

@app.route('/api/faculty', methods=['GET', 'POST'])
def manage_faculty():
    """Get all faculty or add new faculty"""
    if request.method == 'GET':
        result = get_all_documents('faculty')
        return jsonify(result)
    
    elif request.method == 'POST':
        data = request.json
        faculty_id = data.get('id', str(uuid.uuid4()))
        result = add_document('faculty', faculty_id, data)
        return jsonify(result)

@app.route('/api/rooms', methods=['GET', 'POST'])
def manage_rooms():
    """Get all rooms or add new room"""
    if request.method == 'GET':
        result = get_all_documents('rooms')
        return jsonify(result)
    
    elif request.method == 'POST':
        data = request.json
        room_id = data.get('id', str(uuid.uuid4()))
        result = add_document('rooms', room_id, data)
        return jsonify(result)

@app.route('/api/students', methods=['GET', 'POST'])
def manage_students():
    """Get all students or add new student"""
    if request.method == 'GET':
        result = get_all_documents('students')
        return jsonify(result)
    
    elif request.method == 'POST':
        data = request.json
        student_id = data.get('id', str(uuid.uuid4()))
        result = add_document('students', student_id, data)
        return jsonify(result)

@app.route('/api/generate-timetable', methods=['POST'])
def generate_timetable():
    """
    Generate timetable using AI/ML algorithm
    Request body: {
        "program": "B.Ed.",
        "semester": "Semester 1",
        "algorithm": "csp" or "greedy"
    }
    """
    try:
        data = request.json
        program = data.get('program', 'General')
        semester = data.get('semester', 'Current')
        algorithm = data.get('algorithm', 'csp')
        
        courses_result = get_all_documents('courses')
        faculty_result = get_all_documents('faculty')
        rooms_result = get_all_documents('rooms')
        students_result = get_all_documents('students')
        
        if not all([courses_result['success'], faculty_result['success'], 
                   rooms_result['success'], students_result['success']]):
            return jsonify({
                'success': False,
                'message': 'Failed to fetch required data from database'
            }), 500
        
        courses = courses_result['data']
        faculty = faculty_result['data']
        rooms = rooms_result['data']
        students = students_result['data']
        
        # Filter 
        if program != 'General':
            courses = [c for c in courses if c.get('program') == program]
        
        if not courses:
            return jsonify({
                'success': False,
                'message': f'No courses found for program: {program}'
            }), 400
        
        # Program configuration
        program_config = {
            'name': program,
            'semester': semester
        }
        
        # Initialize time
        generator = TimetableGenerator(courses, faculty, rooms, students, program_config)
        
        if algorithm == 'csp':
            result = generator.generate_timetable_csp()
        else:
            result = generator.generate_simple_timetable()
        
        if result['success']:
            validation = generator.validate_timetable(result['timetable'])
            result['validation'] = validation
            
            # Save to Firebase
            timetable_id = str(uuid.uuid4())
            timetable_data = {
                'id': timetable_id,
                'program': program,
                'semester': semester,
                'timetable': result['timetable'],
                'metadata': result['metadata'],
                'validation': validation,
                'created_at': datetime.now().isoformat()
            }
            
            save_result = add_document('timetables', timetable_id, timetable_data)
            
            if save_result['success']:
                result['timetable_id'] = timetable_id
        
        return jsonify(result)
    
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Error generating timetable: {str(e)}'
        }), 500

@app.route('/api/timetable/<timetable_id>', methods=['GET'])
def get_timetable(timetable_id):
    """Get a specific timetable by ID"""
    result = get_document('timetables', timetable_id)
    return jsonify(result)

@app.route('/api/timetables', methods=['GET'])
def get_all_timetables():
    """Get all generated timetables"""
    result = get_all_documents('timetables')
    return jsonify(result)

# Export Endpoints

@app.route('/api/export/pdf/<timetable_id>', methods=['GET'])
def export_pdf(timetable_id):
    """Export timetable to PDF"""
    try:
        result = get_document('timetables', timetable_id)
        
        if not result['success']:
            return jsonify({'success': False, 'message': 'Timetable not found'}), 404
        
        timetable_data = result['data']
        timetable = timetable_data['timetable']
        metadata = timetable_data.get('metadata', {})
        
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), 
                              rightMargin=30, leftMargin=30, 
                              topMargin=30, bottomMargin=18)
        elements = []
        styles = getSampleStyleSheet()
        
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#1a237e'),
            spaceAfter=30,
            alignment=1  
        )
        
        title_text = f"Timetable - {metadata.get('program', 'General')} - {metadata.get('semester', 'Current')}"
        title = Paragraph(title_text, title_style)
        elements.append(title)
        
        days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']
        
        for day in days:
            day_entries = [e for e in timetable if e['day'] == day]
            
            if day_entries:
                day_header = Paragraph(f"<b>{day}</b>", styles['Heading2'])
                elements.append(day_header)
                elements.append(Spacer(1, 12))
                
                table_data = [['Time', 'Course Code', 'Course Name', 'Faculty', 'Room', 'Type']]
                
                for entry in sorted(day_entries, key=lambda x: x['time']):
                    table_data.append([
                        entry['time'],
                        entry['course_code'],
                        entry['course_name'][:30],  
                        entry['faculty_name'][:20],
                        entry['room_number'],
                        entry['type'].capitalize()
                    ])
                
                table = Table(table_data, colWidths=[1.2*inch, 1*inch, 2.5*inch, 1.5*inch, 0.8*inch, 0.8*inch])
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a237e')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('FONTSIZE', (0, 1), (-1, -1), 8),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
                ]))
                
                elements.append(table)
                elements.append(Spacer(1, 20))
        
        footer_text = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | NEP 2020 Compliant"
        footer = Paragraph(footer_text, styles['Normal'])
        elements.append(Spacer(1, 12))
        elements.append(footer)
        
        # Build PDF
        doc.build(elements)
        buffer.seek(0)
        
        return send_file(
            buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f'timetable_{timetable_id}.pdf'
        )
    
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/export/excel/<timetable_id>', methods=['GET'])
def export_excel(timetable_id):
    """Export timetable to Excel"""
    try:
        result = get_document('timetables', timetable_id)
        
        if not result['success']:
            return jsonify({'success': False, 'message': 'Timetable not found'}), 404
        
        timetable_data = result['data']
        timetable = timetable_data['timetable']
        metadata = timetable_data.get('metadata', {})
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Timetable"
        
        ws.merge_cells('A1:G1')
        title_cell = ws['A1']
        title_cell.value = f"Timetable - {metadata.get('program', 'General')} - {metadata.get('semester', 'Current')}"
        title_cell.font = Font(size=16, bold=True, color="1a237e")
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        headers = ['Day', 'Time', 'Course Code', 'Course Name', 'Faculty', 'Room', 'Type']
        header_row = 3
        
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1a237e", end_color="1a237e", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        # Data rows
        days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']
        row = header_row + 1
        
        for day in days:
            day_entries = [e for e in timetable if e['day'] == day]
            
            for entry in sorted(day_entries, key=lambda x: x['time']):
                ws.cell(row=row, column=1, value=entry['day'])
                ws.cell(row=row, column=2, value=entry['time'])
                ws.cell(row=row, column=3, value=entry['course_code'])
                ws.cell(row=row, column=4, value=entry['course_name'])
                ws.cell(row=row, column=5, value=entry['faculty_name'])
                ws.cell(row=row, column=6, value=entry['room_number'])
                ws.cell(row=row, column=7, value=entry['type'].capitalize())
                
                for col in range(1, 8):
                    cell = ws.cell(row=row, column=col)
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                    cell.border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                    
                    if row % 2 == 0:
                        cell.fill = PatternFill(start_color="f0f0f0", end_color="f0f0f0", fill_type="solid")
                
                row += 1
        
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 35
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 10
        ws.column_dimensions['G'].width = 12
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return send_file(
            buffer,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'timetable_{timetable_id}.xlsx'
        )
    
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/validate-data', methods=['POST'])
def validate_data():
    """Validate input data before timetable generation"""
    try:
        data = request.json
        errors = []
        warnings = []
        
        courses_result = get_all_documents('courses')
        faculty_result = get_all_documents('faculty')
        rooms_result = get_all_documents('rooms')
        
        if not courses_result['success'] or len(courses_result.get('data', [])) == 0:
            errors.append('No courses found. Please add courses first.')
        
        if not faculty_result['success'] or len(faculty_result.get('data', [])) == 0:
            errors.append('No faculty found. Please add faculty members first.')
        
        if not rooms_result['success'] or len(rooms_result.get('data', [])) == 0:
            errors.append('No rooms found. Please add rooms first.')
        
        if courses_result['success']:
            courses = courses_result['data']
            faculty = faculty_result.get('data', [])
            
            for course in courses:
                if not course.get('faculty_id'):
                    warnings.append(f"Course '{course.get('name')}' has no assigned faculty")
            
            faculty_course_count = {}
            for course in courses:
                fid = course.get('faculty_id')
                if fid:
                    faculty_course_count[fid] = faculty_course_count.get(fid, 0) + 1
            
            for fid, count in faculty_course_count.items():
                if count > 5:
                    faculty_name = next((f['name'] for f in faculty if f['id'] == fid), 'Unknown')
                    warnings.append(f"Faculty '{faculty_name}' is assigned {count} courses (high workload)")
        
        return jsonify({
            'success': len(errors) == 0,
            'errors': errors,
            'warnings': warnings
        })
    
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)